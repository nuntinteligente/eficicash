"""
App Streamlit: Eficicash — Seção "Folha de Pagamento"
Coloque este arquivo em f:\proopor\app.py e rode com:
    python -m streamlit run f:\proopor\app.py
"""
import streamlit as st
import pandas as pd
import json
import logging
from io import BytesIO
from datetime import date

logging.basicConfig(level=logging.INFO)
logger = logging.getLogger("eficicash")

st.set_page_config(page_title="Eficicash - Folha de Pagamento", layout="wide")

# ---------------- UI: topo e navegação ----------------
st.title("Eficicash — Ferramenta: Folha de Pagamento")
st.markdown("Resultados para orientação. Confirme com sua contabilidade.")
st.markdown(
    "Visite: "
    "[Eficicash (site de contabilidade)]"
    "(https://dossantosmachadoma.wixstudio.com/contabilidade)"
)

# ---------------- Sidebar: navegação e parâmetros ----------------
st.sidebar.title("Navegação")
page = st.sidebar.radio("Seção", ["Dashboard", "Folha de Pagamento", "Lançamentos Contábeis", "Sobre / Export"])

st.sidebar.markdown("---")
st.sidebar.header("Parâmetros padrão (editar se necessário)")
# Padrões ilustrativos; usuário pode editar JSON abaixo
default_inss = [
    {"min": 0.0, "max": 1212.00, "aliquota": 7.5},
    {"min": 1212.01, "max": 2427.35, "aliquota": 9.0},
    {"min": 2427.36, "max": 3641.03, "aliquota": 12.0},
    {"min": 3641.04, "max": 7087.22, "aliquota": 14.0},
]
default_irrf = [
    {"min": 0.0, "max": 1903.98, "aliquota": 0.0, "parcela": 0.0},
    {"min": 1903.99, "max": 2826.65, "aliquota": 7.5, "parcela": 142.80},
    {"min": 2826.66, "max": 3751.05, "aliquota": 15.0, "parcela": 354.80},
    {"min": 3751.06, "max": 4664.68, "aliquota": 22.5, "parcela": 636.13},
    {"min": 4664.69, "max": None, "aliquota": 27.5, "parcela": 869.36},
]
inss_json = st.sidebar.text_area("Faixas INSS (JSON)", value=json.dumps(default_inss, indent=2), height=140)
irrf_json = st.sidebar.text_area("Faixas IRRF (JSON)", value=json.dumps(default_irrf, indent=2), height=160)

deducao_por_dependente = st.sidebar.number_input("Dedução por dependente (R$)", value=189.59, step=0.01, format="%.2f")
fgts_percent = st.sidebar.number_input("FGTS (%)", value=8.0, step=0.1)
inss_patronal_percent = st.sidebar.number_input("INSS patronal (%)", value=20.0, step=0.1)

# ...existing code...
def sobre_tab():
    """Aba 'Sobre / Export' — ajuste o conteúdo conforme necessário."""
    import streamlit as st  # já deve estar importado no topo; mantido por segurança
    st.header("Sobre / Export")
    st.write("Informações sobre o aplicativo. Ajuste este conteúdo conforme necessário.")
    st.markdown("Links e opções de exportação podem ser adicionados aqui.")
# ...existing code...
# ---------------- Utilitários ----------------
def safe_load_json_list(txt, default):
    try:
        parsed = json.loads(txt)
        if isinstance(parsed, list):
            return parsed
    except Exception as e:
        logger.warning("Erro ao analisar JSON: %s", e)
    return default

inss_faixas = safe_load_json_list(inss_json, default_inss)
irrf_faixas = safe_load_json_list(irrf_json, default_irrf)

def calc_valor_hora(salario_base, horas_normais_mes):
    if horas_normais_mes and horas_normais_mes > 0:
        return salario_base / horas_normais_mes
    return 0.0

def calc_proventos(inputs):
    sb = inputs["salario_base"]
    vh = inputs["valor_hora"] if inputs["valor_hora"] is not None else calc_valor_hora(sb, inputs["horas_normais_mes"])
    proventos = []
    proventos.append({"desc": "Salário Base", "base": sb, "aliquota": None, "valor": sb})
    he50 = vh * inputs["horas_extra_50"] * 1.5
    if inputs["horas_extra_50"] > 0:
        proventos.append({"desc": "Horas Extra 50%", "base": f"{vh:.2f} x {inputs['horas_extra_50']}", "aliquota": "50%", "valor": he50})
    he100 = vh * inputs["horas_extra_100"] * 2.0
    if inputs["horas_extra_100"] > 0:
        proventos.append({"desc": "Horas Extra 100%", "base": f"{vh:.2f} x {inputs['horas_extra_100']}", "aliquota": "100%", "valor": he100})
    adicional_noturno_valor = vh * inputs["horas_noturnas"] * (inputs["adicional_noturno_percent"] / 100.0)
    if inputs["horas_noturnas"] > 0:
        proventos.append({"desc": "Adicional Noturno", "base": f"{vh:.2f} x {inputs['horas_noturnas']}", "aliquota": f"{inputs['adicional_noturno_percent']}%", "valor": adicional_noturno_valor})
    if inputs["possui_periculosidade"]:
        perc_val = sb * (inputs["periculosidade_percent"] / 100.0)
        proventos.append({"desc": "Periculosidade", "base": sb, "aliquota": f"{inputs['periculosidade_percent']}%", "valor": perc_val})
    if inputs["possui_insalubridade"]:
        insal_val = sb * (inputs["insalubridade_percent"] / 100.0)
        proventos.append({"desc": "Insalubridade", "base": sb, "aliquota": f"{inputs['insalubridade_percent']}%", "valor": insal_val})
    if inputs["vale_refeicao"] and inputs["vale_refeicao"] > 0:
        proventos.append({"desc": "Vale Refeição (benefício informado)", "base": inputs["vale_refeicao"], "aliquota": None, "valor": inputs["vale_refeicao"]})
    if inputs["outros_proventos"] and inputs["outros_proventos"] > 0:
        proventos.append({"desc": "Outros Proventos", "base": inputs["outros_proventos"], "aliquota": None, "valor": inputs["outros_proventos"]})
    bruto = sum([p["valor"] for p in proventos])
    return {"proventos": proventos, "salario_bruto": bruto, "valor_hora": vh, "he50": he50, "he100": he100}

def calc_inss_progressivo(salario_bruto, faixas):
    parcelas = []
    total = 0.0
    for f in faixas:
        fmin = float(f.get("min", 0.0))
        fmax = None if f.get("max", None) is None else float(f.get("max"))
        ali = float(f.get("aliquota", 0.0)) / 100.0
        if salario_bruto <= fmin:
            base_faixa = 0.0
        else:
            if fmax is None:
                base_faixa = salario_bruto - fmin
            else:
                base_faixa = max(0.0, min(salario_bruto, fmax) - fmin)
        valor = base_faixa * ali
        parcelas.append({"faixa": f"{fmin:.2f} - {('∞' if fmax is None else f'{fmax:.2f}')}", "base": base_faixa, "aliquota": f.get("aliquota", 0.0), "valor": valor})
        total += valor
    return parcelas, total

def calc_irrf(base_irrf, faixas, deducao_dependentes, num_dependentes, pensao):
    ded_depend = num_dependentes * deducao_dependentes
    base = base_irrf - ded_depend - pensao
    base = max(0.0, base)
    aplicada = None
    valor_irrf = 0.0
    for f in faixas:
        fmin = float(f.get("min", 0.0))
        fmax = None if f.get("max", None) is None else float(f.get("max"))
        ali = float(f.get("aliquota", 0.0)) / 100.0
        parc = float(f.get("parcela", 0.0))
        if (base >= fmin) and (fmax is None or base <= fmax):
            valor_irrf = base * ali - parc
            aplicada = f
            break
    valor_irrf = max(0.0, valor_irrf)
    return {"base": base, "irrf": valor_irrf, "faixa_aplicada": aplicada, "deducao_dependentes": ded_depend}

def calc_fgts_provisoes(salario_base, salario_bruto, fgts_percent, inss_patronal_percent):
    fgts = salario_bruto * (fgts_percent / 100.0)
    inss_patronal = salario_bruto * (inss_patronal_percent / 100.0)
    provision_13 = salario_base / 12.0
    provision_ferias = salario_base / 12.0
    provision_ferias_1_3 = provision_ferias / 3.0
    return {
        "fgts": fgts,
        "inss_patronal": inss_patronal,
        "provision_13_mensal": provision_13,
        "provision_ferias_mensal": provision_ferias,
        "provision_ferias_1_3_mensal": provision_ferias_1_3
    }

# --- Tratamento das bibliotecas opcionais (PDF) ---
try:
    from reportlab.lib.pagesizes import A4
    from reportlab.pdfgen import canvas
    PDF_AVAILABLE = True
except Exception:
    PDF_AVAILABLE = False
    logging.getLogger("eficicash").warning("reportlab não encontrado — exportação para PDF desabilitada")

def generate_pdf_bytes(holerite_df, resumo, salario_base, salario_liquido, total_proventos, total_descontos):
    buffer = BytesIO()
    c = canvas.Canvas(buffer, pagesize=A4)
    width, height = A4
    x = 40
    y = height - 40
    c.setFont("Helvetica-Bold", 14)
    c.drawString(x, y, "Holerite - Eficicash")
    c.setFont("Helvetica", 10)
    y -= 20
    c.drawString(x, y, f"Data: {date.today().isoformat()}")
    y -= 20
    c.drawString(x, y, f"Salário Base: R$ {salario_base:.2f}  | Salário Líquido: R$ {salario_liquido:.2f}")
    y -= 30
    c.setFont("Helvetica-Bold", 11)
    c.drawString(x, y, "Proventos:")
    y -= 18
    c.setFont("Helvetica", 10)
    for _, row in holerite_df[holerite_df['tipo']=="Provento"].iterrows():
        if y < 80:
            c.showPage(); y = height - 40
        c.drawString(x, y, f"- {row['descricao']}: R$ {float(row['valor']):.2f}")
        y -= 14
    y -= 10
    c.setFont("Helvetica-Bold", 11)
    c.drawString(x, y, "Descontos:")
    y -= 18
    c.setFont("Helvetica", 10)
    for _, row in holerite_df[holerite_df['tipo']=="Desconto"].iterrows():
        if y < 80:
            c.showPage(); y = height - 40
        c.drawString(x, y, f"- {row['descricao']}: R$ {float(row['valor']):.2f}")
        y -= 14
    y -= 20
    c.setFont("Helvetica-Bold", 11)
    c.drawString(x, y, f"Total Proventos: R$ {total_proventos:.2f}   Total Descontos: R$ {total_descontos:.2f}")
    y -= 16
    c.drawString(x, y, f"Salário Líquido: R$ {salario_liquido:.2f}")
    c.showPage()
    c.save()
    buffer.seek(0)
    return buffer

# --- Adicionar DEFINIÇÃO da função generate_dre (se ainda não existir) ---
def generate_dre(custo_total_empregador, receitas_operacionais=0.0, outras_despesas=0.0):
    """
    Gera DataFrame DRE simplificado:
    Receita Operacional
    - Despesas Operacionais (Despesas com Pessoal, Outras Despesas)
    = Resultado Operacional
    """
    receita = float(receitas_operacionais)
    desp_pessoal = float(custo_total_empregador)
    desp_outras = float(outras_despesas)
    total_despesas = desp_pessoal + desp_outras
    resultado_operacional = receita - total_despesas
    rows = [
        {"Conta": "Receita Operacional", "Valor (R$)": receita, "Fórmula": f"= {receita:.2f}"},
        {"Conta": "Despesas com Pessoal (custo empregador)", "Valor (R$)": desp_pessoal, "Fórmula": f"= custo_total_empregador = {desp_pessoal:.2f}"},
        {"Conta": "Outras Despesas", "Valor (R$)": desp_outras, "Fórmula": f"= {desp_outras:.2f}"},
        {"Conta": "Total Despesas", "Valor (R$)": total_despesas, "Fórmula": f"= {desp_pessoal:.2f} + {desp_outras:.2f} = {total_despesas:.2f}"},
        {"Conta": "Resultado Operacional", "Valor (R$)": resultado_operacional, "Fórmula": f"= {receita:.2f} - {total_despesas:.2f} = {resultado_operacional:.2f}"},
    ]
    return pd.DataFrame(rows)

# ---------------- Função principal da aba Folha de Pagamento ----------------
def folha_pagamento_tab():
    st.header("Folha de Pagamento")
    st.markdown("Preencha os campos obrigatórios. Campos com valores padrão são editáveis.")
    with st.form("form_folha"):
        col1, col2, col3 = st.columns(3)
        with col1:
            salario_base = st.number_input("salario_base (R$ mensais)", min_value=0.0, value=3500.00, step=0.01, format="%.2f")
            dias_uteis_mes = st.number_input("dias_uteis_mes (inteiro)", min_value=1, value=30, step=1)
            horas_normais_mes = st.number_input("horas_normais_mes", min_value=1.0, value=220.0, step=0.5)
            horas_extra_50 = st.number_input("horas_extra_50", min_value=0.0, value=5.0, step=0.5)
            horas_extra_100 = st.number_input("horas_extra_100", min_value=0.0, value=2.0, step=0.5)
            valor_hora_input = st.number_input("valor_hora (0 para calcular automaticamente)", min_value=0.0, value=0.0, step=0.01)
        with col2:
            horas_noturnas = st.number_input("horas_noturnas", min_value=0.0, value=10.0, step=0.5)
            adicional_noturno_percent = st.number_input("adicional_noturno_percent (%)", min_value=0.0, value=20.0, step=0.1)
            possui_periculosidade = st.checkbox("possui_periculosidade", value=False)
            periculosidade_percent = st.number_input("periculosidade_percent (%)", min_value=0.0, value=30.0, step=0.1)
            possui_insalubridade = st.checkbox("possui_insalubridade", value=False)
            insalubridade_percent = st.number_input("insalubridade_percent (%)", min_value=0.0, value=0.0, step=0.1)
            numero_dependentes = st.number_input("numero_dependentes", min_value=0, value=1, step=1)
        with col3:
            pensao_alimenticia = st.number_input("pensao_alimenticia (R$ mensais)", min_value=0.0, value=0.0, step=0.01)
            vt_type = st.selectbox("Vale-transporte (tipo)", ["percentual (6%)", "valor fixo (R$)"])
            if vt_type.startswith("percentual"):
                vale_transporte_percent = st.number_input("vale_transporte_percent (%)", min_value=0.0, value=6.0, step=0.1)
                vale_transporte_valor = None
            else:
                vale_transporte_valor = st.number_input("vale_transporte (R$ mensais)", min_value=0.0, value=0.0, step=0.01)
                vale_transporte_percent = None
            vale_refeicao = st.number_input("vale_refeicao (R$ mensais)", min_value=0.0, value=200.0, step=0.01)
            faltas = st.number_input("faltas (dias no mês)", min_value=0, value=0, step=1)
            atrasos_minutos = st.number_input("atrasos_minutos (minutos totais no mês)", min_value=0, value=0, step=1)
            outros_proventos = st.number_input("outros_proventos (R$)", min_value=0.0, value=0.0, step=0.01)

        calcular = st.form_submit_button("Calcular")

    # somente executa e usa variáveis depois que o usuário clica em Calcular
    if not calcular:
        st.info("Preencha o formulário e clique em 'Calcular' para gerar o holerite e a DRE.")
        return

    # prepara inputs e valida
    inputs = {
        "salario_base": float(salario_base),
        "dias_uteis_mes": int(dias_uteis_mes),
        "horas_normais_mes": float(horas_normais_mes),
        "horas_extra_50": float(horas_extra_50),
        "horas_extra_100": float(horas_extra_100),
        "valor_hora": None if valor_hora_input == 0.0 else float(valor_hora_input),
        "horas_noturnas": float(horas_noturnas),
        "adicional_noturno_percent": float(adicional_noturno_percent),
        "possui_periculosidade": bool(possui_periculosidade),
        "periculosidade_percent": float(periculosidade_percent),
        "possui_insalubridade": bool(possui_insalubridade),
        "insalubridade_percent": float(insalubridade_percent),
        "numero_dependentes": int(numero_dependentes),
        "pensao_alimenticia": float(pensao_alimenticia),
        "vale_transporte_percent": vale_transporte_percent,
        "vale_transporte_valor": vale_transporte_valor,
        "vale_refeicao": float(vale_refeicao),
        "faltas": int(faltas),
        "atrasos_minutos": int(atrasos_minutos),
        "outros_proventos": float(outros_proventos)
    }

    if inputs["salario_base"] < 0 or inputs["horas_normais_mes"] <= 0:
        st.error("salario_base deve ser >= 0 e horas_normais_mes > 0")
        return

    # cálculos (reaproveita funções utilitárias já definidas no arquivo)
    det = calc_proventos(inputs)
    salario_bruto = det["salario_bruto"]

    parcelas_inss, total_inss = calc_inss_progressivo(salario_bruto, inss_faixas)
    irrf_res = calc_irrf(salario_bruto - total_inss, irrf_faixas, deducao_por_dependente, inputs["numero_dependentes"], inputs["pensao_alimenticia"])

    # descontos variáveis
    descontos = []
    vt_desconto = 0.0
    if inputs["vale_transporte_percent"] is not None:
        vt_desconto = inputs["salario_base"] * (inputs["vale_transporte_percent"] / 100.0)
        descontos.append({"desc": "Vale-transporte (desconto %)", "base": inputs["salario_base"], "aliquota": f"{inputs['vale_transporte_percent']}%", "valor": vt_desconto})
    elif inputs["vale_transporte_valor"] is not None:
        vt_desconto = inputs["vale_transporte_valor"]
        descontos.append({"desc": "Vale-transporte (valor fixo)", "base": inputs["vale_transporte_valor"], "aliquota": None, "valor": vt_desconto})

    if inputs["vale_refeicao"] and inputs["vale_refeicao"] > 0:
        descontos.append({"desc": "Vale-refeição (desconto informado)", "base": inputs["vale_refeicao"], "aliquota": None, "valor": inputs["vale_refeicao"]})

    desconto_faltas = (inputs["salario_base"] / inputs["dias_uteis_mes"]) * inputs["faltas"]
    if inputs["faltas"] > 0:
        descontos.append({"desc": "Desconto por faltas", "base": f"{inputs['salario_base']:.2f}/{inputs['dias_uteis_mes']}", "aliquota": None, "valor": desconto_faltas})

    atraso_horas = inputs["atrasos_minutos"] / 60.0
    desconto_atrasos = atraso_horas * det["valor_hora"]
    if inputs["atrasos_minutos"] > 0:
        descontos.append({"desc": "Desconto por atrasos (horas)", "base": f"{atraso_horas:.2f}h x {det['valor_hora']:.2f}", "aliquota": None, "valor": desconto_atrasos})

    if inputs["pensao_alimenticia"] > 0:
        descontos.append({"desc": "Pensão alimentícia", "base": inputs["pensao_alimenticia"], "aliquota": None, "valor": inputs["pensao_alimenticia"]})

    # INSS e IRRF como descontos
    descontos.append({"desc": "INSS (empregado)", "base": salario_bruto, "aliquota": None, "valor": total_inss})
    descontos.append({"desc": "IRRF", "base": irrf_res["base"], "aliquota": float(irrf_res["faixa_aplicada"]["aliquota"]) if irrf_res["faixa_aplicada"] else 0.0, "valor": irrf_res["irrf"]})

    total_descontos = sum([d["valor"] for d in descontos])
    total_proventos = salario_bruto
    salario_liquido = total_proventos - total_descontos

    # provisões e encargos
    prov = calc_fgts_provisoes(inputs["salario_base"], salario_bruto, fgts_percent, inss_patronal_percent)
    custo_total_empregador = salario_bruto + prov["inss_patronal"] + prov["fgts"] + prov["provision_13_mensal"] + prov["provision_ferias_mensal"] + prov["provision_ferias_1_3_mensal"]

    # montar holerite dataframe
    holerite_rows = []
    for p in det["proventos"]:
        holerite_rows.append({"tipo":"Provento","descricao":p["desc"],"base":p["base"],"aliquota":p["aliquota"],"valor":p["valor"]})
    for d in descontos:
        holerite_rows.append({"tipo":"Desconto","descricao":d["desc"],"base":d["base"],"aliquota":d["aliquota"],"valor":d["valor"]})
    holerite_df = pd.DataFrame(holerite_rows)

    # Exibição seções (mantém apresentação passo a passo)
    st.subheader("1) Proventos e salário bruto")
    st.write("Fórmula valor-hora = salario_base / horas_normais_mes")
    st.write(f"Valor-hora: R$ {det['valor_hora']:.2f}")
    st.table(pd.DataFrame(det["proventos"]).rename(columns={"desc":"Descrição","base":"Base","aliquota":"Alíquota","valor":"Valor (R$)"}))
    st.markdown(f"**Salário Bruto:** R$ {salario_bruto:.2f}")

    st.subheader("2) INSS (progressivo por faixa)")
    st.table(pd.DataFrame(parcelas_inss).rename(columns={"faixa":"Faixa","base":"Base (R$)","aliquota":"Alíquota (%)","valor":"Valor (R$)"}))
    st.write(f"Total INSS (empregado): R$ {total_inss:.2f}")

    st.subheader("3) IRRF")
    st.write(f"Base antes de dependentes/pensão: R$ {salario_bruto:.2f} − R$ {total_inss:.2f} = R$ {salario_bruto - total_inss:.2f}")
    st.write(f"IRRF calculado: R$ {irrf_res['irrf']:.2f}")

    st.subheader("4) Descontos")
    st.table(pd.DataFrame(descontos).rename(columns={"desc":"Descrição","base":"Base","aliquota":"Alíquota","valor":"Valor (R$)"}))
    st.write(f"Total de descontos: R$ {total_descontos:.2f}")

    st.subheader("5) FGTS e encargos patronais / provisões")
    st.write(f"FGTS (estimado): R$ {prov['fgts']:.2f}")
    st.write(f"INSS patronal (estimado): R$ {prov['inss_patronal']:.2f}")
    st.write(f"Custo total aproximado do empregador (mensal): R$ {custo_total_empregador:.2f}")

    st.subheader("6) Holerite final")
    st.table(holerite_df.rename(columns={"tipo":"Tipo","descricao":"Descrição","base":"Base","aliquota":"Alíquota","valor":"Valor (R$)"}))
    st.markdown(f"**Total Proventos:** R$ {total_proventos:.2f}  •  **Total Descontos:** R$ {total_descontos:.2f}  •  **Salário Líquido:** R$ {salario_liquido:.2f}")

    # Salva resumo e dataframes no session_state para o Dashboard
    st.session_state['last_holerite_summary'] = {
        "salario_base": float(inputs["salario_base"]),
        "salario_bruto": float(salario_bruto),
        "salario_liquido": float(salario_liquido),
        "total_proventos": float(total_proventos),
        "total_descontos": float(total_descontos),
        "custo_total_empregador": float(custo_total_empregador)
    }
    st.session_state['last_holerite_df'] = holerite_df

    # Entrada opcional para gerar DRE e exibir
    receita_operacional_input = st.number_input("Receita operacional (R$) para DRE (opcional)", min_value=0.0, value=0.0, step=1.0, format="%.2f", key="dre_receita_input")
    outras_despesas_input = st.number_input("Outras despesas (R$) para DRE (opcional)", min_value=0.0, value=0.0, step=1.0, format="%.2f", key="dre_outras_input")

    dre_df = generate_dre(custo_total_empregador, receitas_operacionais=receita_operacional_input, outras_despesas=outras_despesas_input)
    st.session_state['last_dre_df'] = dre_df
    st.subheader("DRE (preliminar) gerada a partir do custo de pessoal")
    st.table(dre_df)

    # Exportação (CSV / XLSX / PDF) - mantém lógica anterior (assume blocos já presentes no arquivo)
    # CSV
    csv_buf = BytesIO()
    holerite_df.to_csv(csv_buf, index=False, sep=";", encoding="utf-8")
    csv_buf.seek(0)
    st.download_button("Baixar CSV do holerite", data=csv_buf, file_name="holerite.csv", mime="text/csv")

    # Excel (.xlsx) com resumo e formatação (usa openpyxl)
    try:
        import openpyxl
        from openpyxl import Workbook
        from openpyxl.utils.dataframe import dataframe_to_rows
        from openpyxl.styles import Font, Alignment, PatternFill, Side, Border, NamedStyle
        excel_buf = BytesIO()

        wb = Workbook()
        ws = wb.active
        ws.title = "Holerite"

        # Cabeçalho / resumo visual
        title_style = Font(size=14, bold=True)
        bold = Font(bold=True)
        money_align = Alignment(horizontal="right")
        header_fill = PatternFill("solid", fgColor="DDDDDD")
        thin = Side(border_style="thin", color="000000")
        border = Border(left=thin, right=thin, top=thin, bottom=thin)

        ws["A1"] = "Eficicash - Holerite"
        ws["A1"].font = title_style

        ws["A3"] = "Salário Base"
        ws["B3"] = float(salario_base)
        ws["A4"] = "Salário Bruto"
        ws["B4"] = float(salario_bruto)
        ws["A5"] = "Total Proventos"
        ws["B5"] = float(total_proventos)
        ws["A6"] = "Total Descontos"
        ws["B6"] = float(total_descontos)
        ws["A7"] = "Salário Líquido"
        ws["B7"] = float(salario_liquido)

        # Formatar células de resumo
        for r in range(3, 8):
            ws[f"A{r}"].font = bold
            ws[f"B{r}"].number_format = '"R$"#,##0.00'
            ws[f"B{r}"].alignment = money_align

        start_row = 9
        # Escrever tabela do holerite com cabeçalho destacada
        rows = list(dataframe_to_rows(holerite_df, index=False, header=True))
        for c, header in enumerate(rows[0], start=1):
            cell = ws.cell(row=start_row, column=c, value=header)
            cell.font = bold
            cell.fill = header_fill
            cell.alignment = Alignment(horizontal="center")

        # Linhas de dados
        for r_idx, row in enumerate(rows[1:], start=start_row + 1):
            for c_idx, value in enumerate(row, start=1):
                cell = ws.cell(row=r_idx, column=c_idx, value=value)
                # aplicar formatação para coluna "valor" (última coluna)
                if holerite_df.columns[-1] == "valor" or holerite_df.columns[-1] == "Valor (R$)":
                    if c_idx == len(row):  # última coluna
                        try:
                            cell.value = float(value)
                            cell.number_format = '"R$"#,##0.00'
                            cell.alignment = money_align
                        except Exception:
                            pass

                # borda leve
                cell.border = border

        # Ajustar largura das colunas
        for col in ws.columns:
            max_length = 0
            col_letter = col[0].column_letter
            for cell in col:
                try:
                    cell_value = str(cell.value)
                    if len(cell_value) > max_length:
                        max_length = len(cell_value)
                except Exception:
                    pass
            adjusted_width = (max_length + 2)
            ws.column_dimensions[col_letter].width = adjusted_width

        # Pequeno rodapé com FGTS / INSS patronal / custo total
        footer_row = start_row + len(rows) + 2
        ws[f"A{footer_row}"] = "FGTS (estimado)"
        ws[f"B{footer_row}"] = float(prov["fgts"])
        ws[f"A{footer_row+1}"] = "INSS patronal (estimado)"
        ws[f"B{footer_row+1}"] = float(prov["inss_patronal"])
        ws[f"A{footer_row+2}"] = "Custo total empregador"
        ws[f"B{footer_row+2}"] = float(custo_total_empregador)
        for r in range(footer_row, footer_row+3):
            ws[f"A{r}"].font = bold
            ws[f"B{r}"].number_format = '"R$"#,##0.00'
            ws[f"B{r}"].alignment = money_align

        wb.save(excel_buf)
        excel_buf.seek(0)

        st.download_button(
            "Baixar Excel (XLSX) do holerite",
            data=excel_buf,
            file_name="holerite.xlsx",
            mime="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet"
        )
    except ImportError:
        st.warning("openpyxl não instalado — instale com: python -m pip install openpyxl")

    # PDF (se disponível)
    if 'PDF_AVAILABLE' in globals() and PDF_AVAILABLE:
        pdf_buf = generate_pdf_bytes(holerite_df, {}, salario_base, salario_liquido, total_proventos, total_descontos)
        st.download_button("Baixar PDF do holerite", data=pdf_buf, file_name="holerite.pdf", mime="application/pdf")
    else:
        st.info("Exportação para PDF desabilitada (instale reportlab para habilitar).")

        # arquivo de ações recomendadas
        acao_text = [
            "1) Revisar benefícios e negociar fornecedor de vale-refeição.",
            "2) Avaliar otimização de jornada e reduzir horas extras.",
            "3) Monitorar DSO e fluxo para evitar atrasos no pagamento da folha."
        ]
        try:
            with open("acao_recomendadas.txt", "w", encoding="utf-8") as f:
                f.write("\n".join(acao_text))
            st.success("Arquivo 'acao_recomendadas.txt' criado no diretório do app.")
        except Exception as e:
            st.warning("Não foi possível gravar arquivo de ações localmente: " + str(e))

# ---------------- Páginas simples para Dashboard e Sobre ----------------
def dashboard_tab():
    st.header("Dashboard (resumo rápido)")
    # chamada para assistente_virtual removida
    st.write("Resumo dos últimos cálculos (se houver).")
    if 'last_holerite_summary' in st.session_state:
        s = st.session_state['last_holerite_summary']
        c1, c2, c3, c4 = st.columns(4)
        c1.metric("Salário Base (R$)", f"{s['salario_base']:.2f}")
        c2.metric("Salário Bruto (R$)", f"{s['salario_bruto']:.2f}")
        c3.metric("Salário Líquido (R$)", f"{s['salario_liquido']:.2f}")
        c4.metric("Custo Empregador (R$)", f"{s['custo_total_empregador']:.2f}")
        st.markdown("**Resumo numérico**")
        st.write(f"Total Proventos: R$ {s['total_proventos']:.2f}  •  Total Descontos: R$ {s['total_descontos']:.2f}")
        if 'last_holerite_df' in st.session_state:
            st.subheader("Último holerite (resumo)")
            st.dataframe(st.session_state['last_holerite_df'].reset_index(drop=True))
        if 'last_dre_df' in st.session_state:
            st.subheader("DRE (último cálculo)")
            st.table(st.session_state['last_dre_df'])
    else:
        st.info("Nenhum cálculo encontrado. Vá para 'Folha de Pagamento' e clique em 'Calcular' para gerar resultados.")

def folha_pagamento_tab():
    st.header("Folha de Pagamento")
    st.markdown("Preencha os campos obrigatórios. Campos com valores padrão são editáveis.")
    with st.form("form_folha"):
        col1, col2, col3 = st.columns(3)
        with col1:
            salario_base = st.number_input("salario_base (R$ mensais)", min_value=0.0, value=3500.00, step=0.01, format="%.2f")
            dias_uteis_mes = st.number_input("dias_uteis_mes (inteiro)", min_value=1, value=30, step=1)
            horas_normais_mes = st.number_input("horas_normais_mes", min_value=1.0, value=220.0, step=0.5)
            horas_extra_50 = st.number_input("horas_extra_50", min_value=0.0, value=5.0, step=0.5)
            horas_extra_100 = st.number_input("horas_extra_100", min_value=0.0, value=2.0, step=0.5)
            valor_hora_input = st.number_input("valor_hora (0 para calcular automaticamente)", min_value=0.0, value=0.0, step=0.01)
        with col2:
            horas_noturnas = st.number_input("horas_noturnas", min_value=0.0, value=10.0, step=0.5)
            adicional_noturno_percent = st.number_input("adicional_noturno_percent (%)", min_value=0.0, value=20.0, step=0.1)
            possui_periculosidade = st.checkbox("possui_periculosidade", value=False)
            periculosidade_percent = st.number_input("periculosidade_percent (%)", min_value=0.0, value=30.0, step=0.1)
            possui_insalubridade = st.checkbox("possui_insalubridade", value=False)
            insalubridade_percent = st.number_input("insalubridade_percent (%)", min_value=0.0, value=0.0, step=0.1)
            numero_dependentes = st.number_input("numero_dependentes", min_value=0, value=1, step=1)
        with col3:
            pensao_alimenticia = st.number_input("pensao_alimenticia (R$ mensais)", min_value=0.0, value=0.0, step=0.01)
            vt_type = st.selectbox("Vale-transporte (tipo)", ["percentual (6%)", "valor fixo (R$)"])
            if vt_type.startswith("percentual"):
                vale_transporte_percent = st.number_input("vale_transporte_percent (%)", min_value=0.0, value=6.0, step=0.1)
                vale_transporte_valor = None
            else:
                vale_transporte_valor = st.number_input("vale_transporte (R$ mensais)", min_value=0.0, value=0.0, step=0.01)
                vale_transporte_percent = None
            vale_refeicao = st.number_input("vale_refeicao (R$ mensais)", min_value=0.0, value=200.0, step=0.01)
            faltas = st.number_input("faltas (dias no mês)", min_value=0, value=0, step=1)
            atrasos_minutos = st.number_input("atrasos_minutos (minutos totais no mês)", min_value=0, value=0, step=1)
            outros_proventos = st.number_input("outros_proventos (R$)", min_value=0.0, value=0.0, step=0.01)

        calcular = st.form_submit_button("Calcular")

    # somente executa e usa variáveis depois que o usuário clica em Calcular
    if not calcular:
        st.info("Preencha o formulário e clique em 'Calcular' para gerar o holerite e a DRE.")
        return

    # prepara inputs e valida
    inputs = {
        "salario_base": float(salario_base),
        "dias_uteis_mes": int(dias_uteis_mes),
        "horas_normais_mes": float(horas_normais_mes),
        "horas_extra_50": float(horas_extra_50),
        "horas_extra_100": float(horas_extra_100),
        "valor_hora": None if valor_hora_input == 0.0 else float(valor_hora_input),
        "horas_noturnas": float(horas_noturnas),
        "adicional_noturno_percent": float(adicional_noturno_percent),
        "possui_periculosidade": bool(possui_periculosidade),
        "periculosidade_percent": float(periculosidade_percent),
        "possui_insalubridade": bool(possui_insalubridade),
        "insalubridade_percent": float(insalubridade_percent),
        "numero_dependentes": int(numero_dependentes),
        "pensao_alimenticia": float(pensao_alimenticia),
        "vale_transporte_percent": vale_transporte_percent,
        "vale_transporte_valor": vale_transporte_valor,
        "vale_refeicao": float(vale_refeicao),
        "faltas": int(faltas),
        "atrasos_minutos": int(atrasos_minutos),
        "outros_proventos": float(outros_proventos)
    }

    if inputs["salario_base"] < 0 or inputs["horas_normais_mes"] <= 0:
        st.error("salario_base deve ser >= 0 e horas_normais_mes > 0")
        return

    # cálculos (reaproveita funções utilitárias já definidas no arquivo)
    det = calc_proventos(inputs)
    salario_bruto = det["salario_bruto"]

    parcelas_inss, total_inss = calc_inss_progressivo(salario_bruto, inss_faixas)
    irrf_res = calc_irrf(salario_bruto - total_inss, irrf_faixas, deducao_por_dependente, inputs["numero_dependentes"], inputs["pensao_alimenticia"])

    # descontos variáveis
    descontos = []
    vt_desconto = 0.0
    if inputs["vale_transporte_percent"] is not None:
        vt_desconto = inputs["salario_base"] * (inputs["vale_transporte_percent"] / 100.0)
        descontos.append({"desc": "Vale-transporte (desconto %)", "base": inputs["salario_base"], "aliquota": f"{inputs['vale_transporte_percent']}%", "valor": vt_desconto})
    elif inputs["vale_transporte_valor"] is not None:
        vt_desconto = inputs["vale_transporte_valor"]
        descontos.append({"desc": "Vale-transporte (valor fixo)", "base": inputs["vale_transporte_valor"], "aliquota": None, "valor": vt_desconto})

    if inputs["vale_refeicao"] and inputs["vale_refeicao"] > 0:
        descontos.append({"desc": "Vale-refeição (desconto informado)", "base": inputs["vale_refeicao"], "aliquota": None, "valor": inputs["vale_refeicao"]})

    desconto_faltas = (inputs["salario_base"] / inputs["dias_uteis_mes"]) * inputs["faltas"]
    if inputs["faltas"] > 0:
        descontos.append({"desc": "Desconto por faltas", "base": f"{inputs['salario_base']:.2f}/{inputs['dias_uteis_mes']}", "aliquota": None, "valor": desconto_faltas})

    atraso_horas = inputs["atrasos_minutos"] / 60.0
    desconto_atrasos = atraso_horas * det["valor_hora"]
    if inputs["atrasos_minutos"] > 0:
        descontos.append({"desc": "Desconto por atrasos (horas)", "base": f"{atraso_horas:.2f}h x {det['valor_hora']:.2f}", "aliquota": None, "valor": desconto_atrasos})

    if inputs["pensao_alimenticia"] > 0:
        descontos.append({"desc": "Pensão alimentícia", "base": inputs["pensao_alimenticia"], "aliquota": None, "valor": inputs["pensao_alimenticia"]})

    # INSS e IRRF como descontos
    descontos.append({"desc": "INSS (empregado)", "base": salario_bruto, "aliquota": None, "valor": total_inss})
    descontos.append({"desc": "IRRF", "base": irrf_res["base"], "aliquota": float(irrf_res["faixa_aplicada"]["aliquota"]) if irrf_res["faixa_aplicada"] else 0.0, "valor": irrf_res["irrf"]})

    total_descontos = sum([d["valor"] for d in descontos])
    total_proventos = salario_bruto
    salario_liquido = total_proventos - total_descontos

    # provisões e encargos
    prov = calc_fgts_provisoes(inputs["salario_base"], salario_bruto, fgts_percent, inss_patronal_percent)
    custo_total_empregador = salario_bruto + prov["inss_patronal"] + prov["fgts"] + prov["provision_13_mensal"] + prov["provision_ferias_mensal"] + prov["provision_ferias_1_3_mensal"]

    # montar holerite dataframe
    holerite_rows = []
    for p in det["proventos"]:
        holerite_rows.append({"tipo":"Provento","descricao":p["desc"],"base":p["base"],"aliquota":p["aliquota"],"valor":p["valor"]})
    for d in descontos:
        holerite_rows.append({"tipo":"Desconto","descricao":d["desc"],"base":d["base"],"aliquota":d["aliquota"],"valor":d["valor"]})
    holerite_df = pd.DataFrame(holerite_rows)

    # Exibição seções (mantém apresentação passo a passo)
    st.subheader("1) Proventos e salário bruto")
    st.write("Fórmula valor-hora = salario_base / horas_normais_mes")
    st.write(f"Valor-hora: R$ {det['valor_hora']:.2f}")
    st.table(pd.DataFrame(det["proventos"]).rename(columns={"desc":"Descrição","base":"Base","aliquota":"Alíquota","valor":"Valor (R$)"}))
    st.markdown(f"**Salário Bruto:** R$ {salario_bruto:.2f}")

    st.subheader("2) INSS (progressivo por faixa)")
    st.table(pd.DataFrame(parcelas_inss).rename(columns={"faixa":"Faixa","base":"Base (R$)","aliquota":"Alíquota (%)","valor":"Valor (R$)"}))
    st.write(f"Total INSS (empregado): R$ {total_inss:.2f}")

    st.subheader("3) IRRF")
    st.write(f"Base antes de dependentes/pensão: R$ {salario_bruto:.2f} − R$ {total_inss:.2f} = R$ {salario_bruto - total_inss:.2f}")
    st.write(f"IRRF calculado: R$ {irrf_res['irrf']:.2f}")

    st.subheader("4) Descontos")
    st.table(pd.DataFrame(descontos).rename(columns={"desc":"Descrição","base":"Base","aliquota":"Alíquota","valor":"Valor (R$)"}))
    st.write(f"Total de descontos: R$ {total_descontos:.2f}")

    st.subheader("5) FGTS e encargos patronais / provisões")
    st.write(f"FGTS (estimado): R$ {prov['fgts']:.2f}")
    st.write(f"INSS patronal (estimado): R$ {prov['inss_patronal']:.2f}")
    st.write(f"Custo total aproximado do empregador (mensal): R$ {custo_total_empregador:.2f}")

    st.subheader("6) Holerite final")
    st.table(holerite_df.rename(columns={"tipo":"Tipo","descricao":"Descrição","base":"Base","aliquota":"Alíquota","valor":"Valor (R$)"}))
    st.markdown(f"**Total Proventos:** R$ {total_proventos:.2f}  •  **Total Descontos:** R$ {total_descontos:.2f}  •  **Salário Líquido:** R$ {salario_liquido:.2f}")

    # Salva resumo e dataframes no session_state para o Dashboard
    st.session_state['last_holerite_summary'] = {
        "salario_base": float(inputs["salario_base"]),
        "salario_bruto": float(salario_bruto),
        "salario_liquido": float(salario_liquido),
        "total_proventos": float(total_proventos),
        "total_descontos": float(total_descontos),
        "custo_total_empregador": float(custo_total_empregador)
    }
    st.session_state['last_holerite_df'] = holerite_df

    # Entrada opcional para gerar DRE e exibir
    receita_operacional_input = st.number_input("Receita operacional (R$) para DRE (opcional)", min_value=0.0, value=0.0, step=1.0, format="%.2f", key="dre_receita_input")
    outras_despesas_input = st.number_input("Outras despesas (R$) para DRE (opcional)", min_value=0.0, value=0.0, step=1.0, format="%.2f", key="dre_outras_input")

    dre_df = generate_dre(custo_total_empregador, receitas_operacionais=receita_operacional_input, outras_despesas=outras_despesas_input)
    st.session_state['last_dre_df'] = dre_df
    st.subheader("DRE (preliminar) gerada a partir do custo de pessoal")
    st.table(dre_df)

    # Exportação (CSV / XLSX / PDF) - mantém lógica anterior (assume blocos já presentes no arquivo)
    # CSV
    csv_buf = BytesIO()
    holerite_df.to_csv(csv_buf, index=False, sep=";", encoding="utf-8")
    csv_buf.seek(0)
    st.download_button("Baixar CSV do holerite", data=csv_buf, file_name="holerite.csv", mime="text/csv")

    # Excel (.xlsx) com resumo e formatação (usa openpyxl)
    try:
        import openpyxl
        from openpyxl import Workbook
        from openpyxl.utils.dataframe import dataframe_to_rows
        from openpyxl.styles import Font, Alignment, PatternFill, Side, Border, NamedStyle
        excel_buf = BytesIO()

        wb = Workbook()
        ws = wb.active
        ws.title = "Holerite"

        # Cabeçalho / resumo visual
        title_style = Font(size=14, bold=True)
        bold = Font(bold=True)
        money_align = Alignment(horizontal="right")
        header_fill = PatternFill("solid", fgColor="DDDDDD")
        thin = Side(border_style="thin", color="000000")
        border = Border(left=thin, right=thin, top=thin, bottom=thin)

        ws["A1"] = "Eficicash - Holerite"
        ws["A1"].font = title_style

        ws["A3"] = "Salário Base"
        ws["B3"] = float(salario_base)
        ws["A4"] = "Salário Bruto"
        ws["B4"] = float(salario_bruto)
        ws["A5"] = "Total Proventos"
        ws["B5"] = float(total_proventos)
        ws["A6"] = "Total Descontos"
        ws["B6"] = float(total_descontos)
        ws["A7"] = "Salário Líquido"
        ws["B7"] = float(salario_liquido)

        # Formatar células de resumo
        for r in range(3, 8):
            ws[f"A{r}"].font = bold
            ws[f"B{r}"].number_format = '"R$"#,##0.00'
            ws[f"B{r}"].alignment = money_align

        start_row = 9
        # Escrever tabela do holerite com cabeçalho destacada
        rows = list(dataframe_to_rows(holerite_df, index=False, header=True))
        for c, header in enumerate(rows[0], start=1):
            cell = ws.cell(row=start_row, column=c, value=header)
            cell.font = bold
            cell.fill = header_fill
            cell.alignment = Alignment(horizontal="center")

        # Linhas de dados
        for r_idx, row in enumerate(rows[1:], start=start_row + 1):
            for c_idx, value in enumerate(row, start=1):
                cell = ws.cell(row=r_idx, column=c_idx, value=value)
                # aplicar formatação para coluna "valor" (última coluna)
                if holerite_df.columns[-1] == "valor" or holerite_df.columns[-1] == "Valor (R$)":
                    if c_idx == len(row):  # última coluna
                        try:
                            cell.value = float(value)
                            cell.number_format = '"R$"#,##0.00'
                            cell.alignment = money_align
                        except Exception:
                            pass

                # borda leve
                cell.border = border

        # Ajustar largura das colunas
        for col in ws.columns:
            max_length = 0
            col_letter = col[0].column_letter
            for cell in col:
                try:
                    cell_value = str(cell.value)
                    if len(cell_value) > max_length:
                        max_length = len(cell_value)
                except Exception:
                    pass
            adjusted_width = (max_length + 2)
            ws.column_dimensions[col_letter].width = adjusted_width

        # Pequeno rodapé com FGTS / INSS patronal / custo total
        footer_row = start_row + len(rows) + 2
        ws[f"A{footer_row}"] = "FGTS (estimado)"
        ws[f"B{footer_row}"] = float(prov["fgts"])
        ws[f"A{footer_row+1}"] = "INSS patronal (estimado)"
        ws[f"B{footer_row+1}"] = float(prov["inss_patronal"])
        ws[f"A{footer_row+2}"] = "Custo total empregador"
        ws[f"B{footer_row+2}"] = float(custo_total_empregador)
        for r in range(footer_row, footer_row+3):
            ws[f"A{r}"].font = bold
            ws[f"B{r}"].number_format = '"R$"#,##0.00'
            ws[f"B{r}"].alignment = money_align

        wb.save(excel_buf)
        excel_buf.seek(0)

        st.download_button(
            "Baixar Excel (XLSX) do holerite",
            data=excel_buf,
            file_name="holerite.xlsx",
            mime="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet"
        )
    except ImportError:
        st.warning("openpyxl não instalado — instale com: python -m pip install openpyxl")

    # PDF (se disponível)
    if 'PDF_AVAILABLE' in globals() and PDF_AVAILABLE:
        pdf_buf = generate_pdf_bytes(holerite_df, {}, salario_base, salario_liquido, total_proventos, total_descontos)
        st.download_button("Baixar PDF do holerite", data=pdf_buf, file_name="holerite.pdf", mime="application/pdf")
    else:
        st.info("Exportação para PDF desabilitada (instale reportlab para habilitar).")

        # arquivo de ações recomendadas
        acao_text = [
            "1) Revisar benefícios e negociar fornecedor de vale-refeição.",
            "2) Avaliar otimização de jornada e reduzir horas extras.",
            "3) Monitorar DSO e fluxo para evitar atrasos no pagamento da folha."
        ]
        try:
            with open("acao_recomendadas.txt", "w", encoding="utf-8") as f:
                f.write("\n".join(acao_text))
            st.success("Arquivo 'acao_recomendadas.txt' criado no diretório do app.")
        except Exception as e:
            st.warning("Não foi possível gravar arquivo de ações localmente: " + str(e))

def lancamentos_contabeis_tab():
    st.header("Lançamentos Contábeis")
    # chamada para assistente_virtual removida
    """
    Seção de Lançamentos Contábeis em Português com exportação para Excel.
    Armazena em st.session_state['lancamentos_df'] com colunas em português.
    """
    import pandas as pd
    from io import BytesIO
    from datetime import date
    try:
        from openpyxl import Workbook
        from openpyxl.styles import Font, Alignment, PatternFill, Border, Side, numbers
    except Exception:
        Workbook = None

    st.markdown("Registre lançamentos e exporte para Excel. Campos: data, descrição, valor, tipo, conta, categoria, data_vencimento, data_recebimento, cliente, fornecedor, centro_custo.")

    # colunas em português
    colunas = ["data", "descrição", "valor", "tipo", "conta", "categoria", "data_vencimento", "data_recebimento", "cliente", "fornecedor", "centro_custo"]
    if 'lancamentos_df' not in st.session_state:
        st.session_state['lancamentos_df'] = pd.DataFrame(columns=colunas)

    # Formulário de inclusão
    with st.form("form_lancamento", clear_on_submit=True):
        c1, c2, c3 = st.columns([1,3,1])
        with c1:
            f_data = st.date_input("Data", value=date.today())
            f_venc = st.date_input("Data de vencimento", value=date.today())
        with c2:
            f_desc = st.text_input("Descrição")
            f_cliente = st.text_input("Cliente")
            f_fornecedor = st.text_input("Fornecedor")
        with c3:
            f_valor = st.number_input("Valor (R$)", value=0.0, format="%.2f", step=0.01)
            f_tipo = st.selectbox("Tipo", ["despesa", "receita", "transferência"])
            f_categoria = st.selectbox("Categoria", ["Gasto", "Ganho", "Outros"])
        f_conta = st.text_input("Conta / Histórico")
        f_recebido = st.date_input("Data de recebimento", value=date.today())
        f_centro = st.text_input("Centro de custo")
        adicionar = st.form_submit_button("Adicionar lançamento")

    if adicionar:
        nova = {
            "data": f_data.isoformat(),
            "descrição": f_desc,
            "valor": float(f_valor),
            "tipo": f_tipo,
            "conta": f_conta,
            "categoria": f_categoria,
            "data_vencimento": f_venc.isoformat(),
            "data_recebimento": f_recebido.isoformat(),
            "cliente": f_cliente,
            "fornecedor": f_fornecedor,
            "centro_custo": f_centro
        }
        df = st.session_state['lancamentos_df']
        st.session_state['lancamentos_df'] = pd.concat([df, pd.DataFrame([nova])], ignore_index=True)
        st.success("Lançamento adicionado.")

    # Exibição e ações
    st.subheader("Lançamentos registrados")
    df_display = st.session_state['lancamentos_df'].copy()
    if df_display.empty:
        st.info("Nenhum lançamento registrado. Use o formulário acima.")
    else:
        total_gasto = df_display.loc[df_display['categoria'] == "Gasto", 'valor'].sum()
        total_ganho = df_display.loc[df_display['categoria'] == "Ganho", 'valor'].sum()
        total_receita = df_display.loc[df_display['tipo'] == "receita", 'valor'].sum()
        total_despesa = df_display.loc[df_display['tipo'] == "despesa", 'valor'].sum()
        a1, a2, a3, a4 = st.columns(4)
        a1.metric("Total Gasto (R$)", f"{total_gasto:.2f}")
        a2.metric("Total Ganho (R$)", f"{total_ganho:.2f}")
        a3.metric("Total Receita (R$)", f"{total_receita:.2f}")
        a4.metric("Total Despesa (R$)", f"{total_despesa:.2f}")

        st.dataframe(df_display.reset_index(drop=True).style.format({"valor": "{:,.2f}"}), height=320)

        # Remover por índice
        with st.expander("Remover lançamento por índice"):
            idx = st.number_input("Índice (linha) a remover", min_value=0, max_value=max(0, len(df_display)-1), step=1, value=0)
            if st.button("Remover índice selecionado"):
                st.session_state['lancamentos_df'] = df_display.drop(index=idx).reset_index(drop=True)
                st.success(f"Lançamento na linha {idx} removido.")
                st.experimental_rerun()

        # Exportações
        e1, e2, e3 = st.columns(3)
        # CSV
        csv_buf = BytesIO()
        df_display.to_csv(csv_buf, index=False, sep=";", encoding="utf-8")
        csv_buf.seek(0)
        e1.download_button("Baixar CSV", data=csv_buf, file_name="lancamentos.csv", mime="text/csv")

        # XLSX com formatação em português se openpyxl disponível
        if Workbook is not None:
            wb = Workbook()
            ws = wb.active
            ws.title = "Lançamentos"
            header_font = Font(bold=True, color="FFFFFF")
            header_fill = PatternFill("solid", fgColor="1F4E78")
            thin = Side(border_style="thin", color="000000")
            border = Border(left=thin, right=thin, top=thin, bottom=thin)
            # cabeçalho em português
            for col_idx, h in enumerate(df_display.columns.tolist(), start=1):
                cell = ws.cell(row=1, column=col_idx, value=h)
                cell.font = header_font
                cell.fill = header_fill
                cell.alignment = Alignment(horizontal="center")
                cell.border = border
            # linhas
            for r_idx, (_, row) in enumerate(df_display.iterrows(), start=2):
                for c_idx, col in enumerate(df_display.columns.tolist(), start=1):
                    val = row[col]
                    cell = ws.cell(row=r_idx, column=c_idx, value=val)
                    cell.border = border
                    if col == "valor":
                        try:
                            cell.value = float(val)
                            cell.number_format = numbers.FORMAT_NUMBER_00
                            if (row["categoria"] == "Gasto") or (row["tipo"] == "despesa"):
                                cell.fill = PatternFill("solid", fgColor="FDE9E9")
                            elif (row["categoria"] == "Ganho") or (row["tipo"] == "receita"):
                                cell.fill = PatternFill("solid", fgColor="E7F9EE")
                        except Exception:
                            pass
                    if col in ("data", "data_vencimento", "data_recebimento"):
                        cell.alignment = Alignment(horizontal="center")
            # ajustar largura
            for col_cells in ws.columns:
                max_len = 0
                col_letter = col_cells[0].column_letter
                for cell in col_cells:
                    try:
                        l = len(str(cell.value))
                        if l > max_len:
                            max_len = l
                    except Exception:
                        pass
                ws.column_dimensions[col_letter].width = max(10, min(50, max_len + 2))
            xlsx_buf = BytesIO()
            wb.save(xlsx_buf)
            xlsx_buf.seek(0)
            e2.download_button("Baixar XLSX", data=xlsx_buf, file_name="lancamentos.xlsx", mime="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet")
        else:
            e2.warning("openpyxl não instalado — execute: python -m pip install openpyxl")

        # Limpar tudo
        if e3.button("Limpar lançamentos"):
            st.session_state['lancamentos_df'] = pd.DataFrame(columns=colunas)
            st.success("Lançamentos removidos.")
# ---------------- Roteamento ----------------
# roteamento principal (substitua/garanta que exista apenas um bloco de roteamento)
if page == "Dashboard":
    dashboard_tab()
elif page == "Folha de Pagamento":
    folha_pagamento_tab()
elif page == "Lançamentos Contábeis":
    lancamentos_contabeis_tab()
else:
    sobre_tab()
